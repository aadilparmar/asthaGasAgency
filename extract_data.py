"""
Comprehensive data extraction from SALARY 2024-25 (4).xlsx
Reads every month sheet (April-24 through MARCH 25), extracts all employee data,
daily deliveries, loan balances, deductions, office staff salaries, and outputs
structured JSON.
"""

import json
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, date

EXCEL_FILE = "SALARY 2024-25 (4).xlsx"
OUTPUT_FILE = "extracted_data.json"

# Month sheets in order, with (month_number, year)
SHEET_CONFIG = [
    ("April-24",  4, 2024),
    ("May-24",    5, 2024),
    ("Jun-24",    6, 2024),
    ("July-24",   7, 2024),
    ("Aug-24",    8, 2024),
    ("Sep-24",    9, 2024),
    ("Oct-24",   10, 2024),
    ("Nov-24",   11, 2024),
    ("Dec-24",   12, 2024),
    ("Jan-25",    1, 2025),
    ("Feb-25",    2, 2025),
    ("MARCH 25",  3, 2025),
]

# Columns that are "system" columns (not real delivery employees)
SYSTEM_LABELS = {"date", "total", "eztef.", "alh,ef.", "~:tdef.", "s],", "gfd"}


def safe_num(val):
    """Convert a cell value to a number, returning 0 for None/empty/whitespace."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if s == "" or s == " ":
        return 0
    try:
        return float(s)
    except ValueError:
        return 0


def safe_str(val):
    """Convert cell value to stripped string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def find_row_by_label(ws, label_lower, col_a=True, col_b=False, start=35, end=55):
    """Find a row where col A (or B) contains the given lowercase substring."""
    for row in range(start, min(end, ws.max_row + 1)):
        if col_a:
            a = ws.cell(row=row, column=1).value
            if a and label_lower in str(a).strip().lower():
                return row
        if col_b:
            b = ws.cell(row=row, column=2).value
            if b and label_lower in str(b).strip().lower():
                return row
    return None


def extract_sheet(wb, sheet_name, month_num, year):
    """Extract all data from a single month sheet."""
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Processing: {sheet_name} (month={month_num}, year={year})")
    print(f"  Dimensions: rows={ws.max_row}, cols={ws.max_column}")

    # --- Determine row offsets ---
    # Most sheets: names on row 6, data starts row 7
    # MARCH 25: names on row 5, data starts row 6
    # Detect by checking if row 5 col B has "Date"
    name_row = 6
    data_start_row = 7
    if safe_str(ws.cell(row=5, column=2).value) and "date" in str(ws.cell(row=5, column=2).value).strip().lower():
        name_row = 5
        data_start_row = 6

    # Loan rows are always relative: loan_opening is name_row - 3, additional is name_row - 2, nett is name_row - 1
    loan_opening_row = name_row - 3
    loan_additional_row = name_row - 2
    nett_loan_row = name_row - 1

    print(f"  name_row={name_row}, data_start_row={data_start_row}")
    print(f"  loan rows: opening={loan_opening_row}, additional={loan_additional_row}, nett={nett_loan_row}")

    # --- Find summary rows dynamically ---
    total_deliver_row = find_row_by_label(ws, "total cly", col_a=True)
    if total_deliver_row is None:
        total_deliver_row = find_row_by_label(ws, "total cy", col_a=True)
    rate_row = find_row_by_label(ws, "deliver rate", col_a=True)
    total_salary_row = find_row_by_label(ws, "total salary", col_a=True)

    # Deduction rows: PF, Lone Instalment, UPAD 1, UAPD 15, UPAD OTHER
    pf_row = find_row_by_label(ws, "pf", col_b=True, start=total_salary_row or 38)
    loan_inst_row = find_row_by_label(ws, "lone instalment", col_b=True, start=pf_row or 38) or \
                    find_row_by_label(ws, "loan instalment", col_b=True, start=pf_row or 38)
    upad1_row = find_row_by_label(ws, "upad 1", col_b=True, start=loan_inst_row or 40)
    upad15_row = find_row_by_label(ws, "uapd 15", col_b=True, start=upad1_row or 41) or \
                 find_row_by_label(ws, "upad 15", col_b=True, start=upad1_row or 41)
    upad_other_row = find_row_by_label(ws, "upad other", col_b=True, start=upad15_row or 42)
    total_deduction_row = find_row_by_label(ws, "total deduction", col_a=True, start=upad_other_row or 43)
    net_payable_row = find_row_by_label(ws, "nate payable", col_a=True, start=total_deduction_row or 44) or \
                      find_row_by_label(ws, "net payable", col_a=True, start=total_deduction_row or 44)
    loan_carry_row = find_row_by_label(ws, "lone cary", col_a=True, start=net_payable_row or 45) or \
                     find_row_by_label(ws, "loan carry", col_a=True, start=net_payable_row or 45)

    print(f"  Summary rows: total_deliver={total_deliver_row}, rate={rate_row}, total_salary={total_salary_row}")
    print(f"  Deduction rows: PF={pf_row}, Loan_Inst={loan_inst_row}, UPAD1={upad1_row}, UPAD15={upad15_row}, UPAD_OTHER={upad_other_row}")
    print(f"  total_deduction={total_deduction_row}, net_payable={net_payable_row}, loan_carry={loan_carry_row}")

    # --- Parse column names from name_row ---
    all_names = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=name_row, column=col).value
        if v is not None:
            all_names[col] = str(v).strip()

    # --- Identify delivery employees (cols C to O typically, before first "Total" at P) ---
    # Find first Total column (delivery total at P=16)
    first_total_col = None
    for col in sorted(all_names.keys()):
        if col >= 3 and all_names[col].lower() == "total":
            first_total_col = col
            break

    # Find second Total column (grand total at U=21)
    second_total_col = None
    for col in sorted(all_names.keys()):
        if col > first_total_col and all_names[col].lower() == "total":
            second_total_col = col
            break

    # Delivery employee columns: from col 3 to first_total_col - 1
    delivery_cols = []
    for col in range(3, first_total_col):
        name = all_names.get(col)
        if name and name.lower() not in SYSTEM_LABELS:
            delivery_cols.append((col, name))

    print(f"  Delivery employees ({len(delivery_cols)}): {[(get_column_letter(c), n) for c, n in delivery_cols]}")

    # --- Office staff columns ---
    # After second Total, find office staff section
    # Office section starts after a gap (usually col W or X)
    # Find the "total" (lowercase) column that ends the office section
    office_total_col = None
    office_cols = []

    # Scan columns after second_total_col for office names
    # Office names are between second_total_col+1 and the next "total" column
    if second_total_col:
        for col in range(second_total_col + 1, ws.max_column + 1):
            name = all_names.get(col)
            if name:
                if name.lower() == "total":
                    office_total_col = col
                    break
                # Skip system labels and "Attendance" label
                if name.lower() not in SYSTEM_LABELS and name.lower() != "attendance":
                    office_cols.append((col, name))

    print(f"  Office employees ({len(office_cols)}): {[(get_column_letter(c), n) for c, n in office_cols]}")
    print(f"  Office total col: {get_column_letter(office_total_col) if office_total_col else None}")

    # --- Determine daily data range ---
    # Data goes from data_start_row to the row before total_deliver_row
    # But some sheets have a "NEW" row right before the summary
    data_end_row = (total_deliver_row - 1) if total_deliver_row else 37
    # Skip the NEW row if present
    if ws.cell(row=data_end_row, column=2).value and str(ws.cell(row=data_end_row, column=2).value).strip().upper() == "NEW":
        data_end_row -= 1

    print(f"  Daily data range: rows {data_start_row} to {data_end_row}")

    # --- Extract delivery employees ---
    delivery_employees = []
    for col, name in delivery_cols:
        col_letter = get_column_letter(col)

        # Loan balances
        loan_opening = safe_num(ws.cell(row=loan_opening_row, column=col).value)
        loan_additional = safe_num(ws.cell(row=loan_additional_row, column=col).value)
        nett_loan = safe_num(ws.cell(row=nett_loan_row, column=col).value)

        # Daily deliveries
        daily = {}
        for row in range(data_start_row, data_end_row + 1):
            date_val = ws.cell(row=row, column=2).value  # Col B = date
            day_num = None
            if isinstance(date_val, datetime):
                day_num = date_val.day
            elif isinstance(date_val, date):
                day_num = date_val.day
            elif date_val:
                # Try parsing string date
                try:
                    d = datetime.strptime(str(date_val).strip(), "%Y-%m-%d %H:%M:%S")
                    day_num = d.day
                except:
                    pass

            if day_num is not None:
                count = safe_num(ws.cell(row=row, column=col).value)
                daily[str(day_num)] = count

        # Total deliveries
        total_deliveries = safe_num(ws.cell(row=total_deliver_row, column=col).value) if total_deliver_row else 0

        # Delivery rate
        rate = safe_num(ws.cell(row=rate_row, column=col).value) if rate_row else 0

        # Total salary
        total_salary = safe_num(ws.cell(row=total_salary_row, column=col).value) if total_salary_row else 0

        # Deductions
        pf = safe_num(ws.cell(row=pf_row, column=col).value) if pf_row else 0
        loan_inst = safe_num(ws.cell(row=loan_inst_row, column=col).value) if loan_inst_row else 0
        upad1 = safe_num(ws.cell(row=upad1_row, column=col).value) if upad1_row else 0
        upad15 = safe_num(ws.cell(row=upad15_row, column=col).value) if upad15_row else 0
        upad_other = safe_num(ws.cell(row=upad_other_row, column=col).value) if upad_other_row else 0
        total_deduction = safe_num(ws.cell(row=total_deduction_row, column=col).value) if total_deduction_row else 0
        net_payable = safe_num(ws.cell(row=net_payable_row, column=col).value) if net_payable_row else 0
        loan_carry = safe_num(ws.cell(row=loan_carry_row, column=col).value) if loan_carry_row else 0

        emp = {
            "name": name,
            "column": col_letter,
            "column_index": col,
            "loan_opening": loan_opening,
            "loan_additional": loan_additional,
            "nett_loan": nett_loan,
            "daily_deliveries": daily,
            "total_deliveries": total_deliveries,
            "rate": rate,
            "total_salary": total_salary,
            "deductions": {
                "pf": pf,
                "loan_instalment": loan_inst,
                "upad1": upad1,
                "upad15": upad15,
                "upad_other": upad_other,
            },
            "total_deduction": total_deduction,
            "net_payable": net_payable,
            "loan_carry_forward": loan_carry,
        }
        delivery_employees.append(emp)

        # Debug for first 2 employees
        if len(delivery_employees) <= 2:
            print(f"    DEL [{col_letter}] {name}: deliveries={total_deliveries}, rate={rate}, salary={total_salary}, "
                  f"loan_open={loan_opening}, loan_add={loan_additional}, loan_inst={loan_inst}, "
                  f"net_pay={net_payable}, days_with_data={len([d for d in daily.values() if d > 0])}")

    # --- Extract office employees ---
    office_employees = []
    for col, name in office_cols:
        col_letter = get_column_letter(col)

        # Loan balances
        loan_opening = safe_num(ws.cell(row=loan_opening_row, column=col).value)
        loan_additional = safe_num(ws.cell(row=loan_additional_row, column=col).value)
        nett_loan = safe_num(ws.cell(row=nett_loan_row, column=col).value)

        # Office staff: fixed salary appears in the total_deliver_row for office columns
        # (This is the row where delivery guys have total cylinder count, but for office it has salary)
        # Some sheets have it in the row just before total_deliver_row too (the last daily-data row area)
        fixed_salary = safe_num(ws.cell(row=total_deliver_row, column=col).value) if total_deliver_row else 0

        # Total salary from total_salary_row
        total_salary = safe_num(ws.cell(row=total_salary_row, column=col).value) if total_salary_row else 0

        # If fixed_salary is 0, use total_salary as the fixed salary
        if fixed_salary == 0 and total_salary > 0:
            fixed_salary = total_salary

        # Also collect any daily attendance data if present (some office staff have counts in daily rows)
        daily_data = {}
        for row in range(data_start_row, data_end_row + 1):
            date_val = ws.cell(row=row, column=2).value
            day_num = None
            if isinstance(date_val, datetime):
                day_num = date_val.day
            elif isinstance(date_val, date):
                day_num = date_val.day
            elif date_val:
                try:
                    d = datetime.strptime(str(date_val).strip(), "%Y-%m-%d %H:%M:%S")
                    day_num = d.day
                except:
                    pass

            if day_num is not None:
                count = safe_num(ws.cell(row=row, column=col).value)
                if count != 0:
                    daily_data[str(day_num)] = count

        # Delivery rate for office cols that have one (rate-based office workers)
        rate = safe_num(ws.cell(row=rate_row, column=col).value) if rate_row else 0

        # Deductions
        pf = safe_num(ws.cell(row=pf_row, column=col).value) if pf_row else 0
        loan_inst = safe_num(ws.cell(row=loan_inst_row, column=col).value) if loan_inst_row else 0
        upad1 = safe_num(ws.cell(row=upad1_row, column=col).value) if upad1_row else 0
        upad15 = safe_num(ws.cell(row=upad15_row, column=col).value) if upad15_row else 0
        upad_other = safe_num(ws.cell(row=upad_other_row, column=col).value) if upad_other_row else 0
        total_deduction = safe_num(ws.cell(row=total_deduction_row, column=col).value) if total_deduction_row else 0
        net_payable = safe_num(ws.cell(row=net_payable_row, column=col).value) if net_payable_row else 0
        loan_carry = safe_num(ws.cell(row=loan_carry_row, column=col).value) if loan_carry_row else 0

        emp = {
            "name": name,
            "column": col_letter,
            "column_index": col,
            "fixed_salary": fixed_salary,
            "rate": rate,
            "loan_opening": loan_opening,
            "loan_additional": loan_additional,
            "nett_loan": nett_loan,
            "daily_data": daily_data if daily_data else {},
            "total_salary": total_salary,
            "deductions": {
                "pf": pf,
                "loan_instalment": loan_inst,
                "upad1": upad1,
                "upad15": upad15,
                "upad_other": upad_other,
            },
            "total_deduction": total_deduction,
            "net_payable": net_payable,
            "loan_carry_forward": loan_carry,
        }
        office_employees.append(emp)

        if len(office_employees) <= 2:
            print(f"    OFF [{col_letter}] {name}: fixed_salary={fixed_salary}, total_salary={total_salary}, "
                  f"loan_open={loan_opening}, loan_inst={loan_inst}, net_pay={net_payable}")

    # --- Grand totals ---
    # Delivery grand total from column U (second_total_col)
    delivery_grand_total_salary = safe_num(ws.cell(row=total_salary_row, column=second_total_col).value) if total_salary_row and second_total_col else 0
    delivery_grand_total_deduction = safe_num(ws.cell(row=total_deduction_row, column=second_total_col).value) if total_deduction_row and second_total_col else 0
    delivery_grand_net_payable = safe_num(ws.cell(row=net_payable_row, column=second_total_col).value) if net_payable_row and second_total_col else 0

    # Office grand total from office_total_col
    office_grand_total_salary = safe_num(ws.cell(row=total_salary_row, column=office_total_col).value) if total_salary_row and office_total_col else 0
    office_grand_total_deduction = safe_num(ws.cell(row=total_deduction_row, column=office_total_col).value) if total_deduction_row and office_total_col else 0
    office_grand_net_payable = safe_num(ws.cell(row=net_payable_row, column=office_total_col).value) if net_payable_row and office_total_col else 0

    print(f"  TOTALS: delivery_salary={delivery_grand_total_salary}, delivery_deduction={delivery_grand_total_deduction}, "
          f"delivery_net={delivery_grand_net_payable}")
    print(f"  TOTALS: office_salary={office_grand_total_salary}, office_deduction={office_grand_total_deduction}, "
          f"office_net={office_grand_net_payable}")

    return {
        "sheet": sheet_name,
        "month": month_num,
        "year": year,
        "structure": {
            "name_row": name_row,
            "data_start_row": data_start_row,
            "data_end_row": data_end_row,
            "total_deliver_row": total_deliver_row,
            "rate_row": rate_row,
            "total_salary_row": total_salary_row,
            "pf_row": pf_row,
            "loan_instalment_row": loan_inst_row,
            "upad1_row": upad1_row,
            "upad15_row": upad15_row,
            "upad_other_row": upad_other_row,
            "total_deduction_row": total_deduction_row,
            "net_payable_row": net_payable_row,
            "loan_carry_forward_row": loan_carry_row,
        },
        "delivery_employees": delivery_employees,
        "office_employees": office_employees,
        "grand_totals": {
            "delivery": {
                "total_salary": delivery_grand_total_salary,
                "total_deduction": delivery_grand_total_deduction,
                "net_payable": delivery_grand_net_payable,
            },
            "office": {
                "total_salary": office_grand_total_salary,
                "total_deduction": office_grand_total_deduction,
                "net_payable": office_grand_net_payable,
            },
        },
    }


def main():
    print(f"Loading workbook: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")

    all_months = []
    for sheet_name, month_num, year in SHEET_CONFIG:
        if sheet_name in wb.sheetnames:
            data = extract_sheet(wb, sheet_name, month_num, year)
            all_months.append(data)
        else:
            print(f"\nWARNING: Sheet '{sheet_name}' not found, skipping.")

    result = {"months": all_months}

    # Write JSON
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False, default=str)

    print(f"\n{'='*60}")
    print(f"JSON written to: {OUTPUT_FILE}")
    print(f"Total months extracted: {len(all_months)}")

    # Summary
    for m in all_months:
        del_count = len(m["delivery_employees"])
        off_count = len(m["office_employees"])
        del_names = [e["name"] for e in m["delivery_employees"]]
        off_names = [e["name"] for e in m["office_employees"]]
        print(f"  {m['sheet']}: {del_count} delivery ({', '.join(del_names)}), "
              f"{off_count} office ({', '.join(off_names)})")


if __name__ == "__main__":
    main()
